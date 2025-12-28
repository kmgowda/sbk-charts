#!/usr/local/bin/python3
# Copyright (c) KMG. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
##
from typing import final

from src.charts import constants
from src.charts.multicharts import SbkMultiCharts
from src.custom_ai.hugging_face import HuggingFace
from openpyxl.styles import Font, Border, Side, Alignment   
import textwrap
from src.stat.storage import StorageStat


warning_msg = ("The AI may hallucinate !."
               " The Summary generated  by generative AI models may not be sufficient and accurate."
               " Its recommended to  analyze the graphs along with generated summary.")

class SbkAI(SbkMultiCharts):
    def __init__(self, version, file):
        super().__init__(version, file)
        self.ai = HuggingFace()

    @final
    def get_columns_values(self, ws):
        columns = self.get_columns_from_worksheet(ws)
        ret = dict()
        for col_name, col_idx in columns.items():
            if col_name not in [constants.ID, constants.HEADER, constants.TYPE,
                                constants.STORAGE, constants.ACTION, constants.LATENCY_TIME_UNIT]:
                values = []
                for row in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=col_idx).value
                    values.append(cell_value)
                ret[col_name] = values
        return ret

    def get_storage_stats(self):
        stats = list()
        for name in self.wb.sheetnames:
            if self.is_r_num_sheet(name):
                ws = self.wb[name]
                storage = self.get_storage_name(ws)
                timeunit = self.get_time_unit(ws)
                action = self.get_action_name(ws)
                # get all the columns of R<Count>
                regular = self.get_columns_values(ws)
                # get all the columns of T<Count>
                t_name = self.get_t_num_sheet_name(name)
                ws = self.wb[t_name]
                total = self.get_columns_values(ws)
                stats.append(StorageStat(storage, timeunit, action, regular, total))
        return stats

    def create_summary_sheet(self):
        sheet = super().create_summary_sheet()
        if sheet is None:
            print("Warning: Could not create summary sheet")
            return None

        self.ai.set_storage_stats(self.get_storage_stats())
        throughput_status, throughput_analysis = self.ai.get_throughput_analysis()
        if not throughput_status:
            print(throughput_analysis)
            return sheet

        latency_status, latency_analysis = self.ai.get_latency_analysis()
        if not latency_status:
            print(latency_analysis)
            return sheet

        # Add analysis text to the summary sheet
        try:
            # Set column width to fit 80 characters
            # Using a larger multiplier to ensure 120 characters fit comfortably
            sheet.column_dimensions['H'].width = 120 * 0.90  # Increased from 0.14 to 0.20 for better fit

            # Find the next available row after existing content
            max_row = sheet.max_row + 3  # Add some spacing
            # Add AI Warning
            warn_cell = sheet.cell(row=max_row, column=8)
            warn_cell.value = warning_msg
            warn_cell.font = Font(size=16, bold=True, color="FFFF0000")
            warn_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Calculate required row height for warning message
            warning_wrapped_lines = []
            for line in warning_msg.split('\n'):
                warning_wrapped_lines.extend(textwrap.wrap(line, width=120))

            # Set row height for latency analysis section
            warn_row_height = max(35, len(warning_wrapped_lines) * 35)
            sheet.row_dimensions[max_row].height = warn_row_height

            max_row = sheet.max_row + 2
            # Add a title for the analysis section
            title_cell = sheet.cell(row=max_row, column=7)
            title_cell.value = "AI Performance Analysis"
            title_cell.font = Font(size=18, bold=True, color="FF0000")  # Red, bold, 18pt

            # Add a title description for the analysis section
            dec_cell = sheet.cell(row=max_row, column=8)
            dec_cell.value = self.ai.get_model_description()
            dec_cell.font = Font(size=16, color="00CF00")

            # Add the Throughput analysis section
            cell = sheet.cell(row=max_row + 2, column=7)
            cell.value = "Throughput Analysis"
            cell.font = Font(size=16, bold=True, color = "EE00FF")

            # Add the analysis text with word wrap
            cell = sheet.cell(row=max_row + 2, column=8)
            cell.value = throughput_analysis
            cell.font = Font(size=14, color = "FF800000")
            cell.border = Border(left=Side(style='thin'),
                                   right=Side(style='thin'),
                                   top=Side(style='thin'),
                                   bottom=Side(style='thin'))

            # Enable text wrapping and set row height to auto-adjust
            cell.alignment = Alignment(wrap_text=True, vertical='top')
                
            # Calculate required row height based on text length and wrap at 80 characters
            wrapped_lines = []
            for line in throughput_analysis.split('\n'):
                wrapped_lines.extend(textwrap.wrap(line, width=120))

            # Set row height (20 points per line, minimum 20)
            row_height = max(25, len(wrapped_lines) * 25)
            sheet.row_dimensions[max_row + 2].height = row_height

            # Add the Latency analysis section
            latency_row = sheet.max_row + 1
            
            # Add the Latency analysis section header
            cell = sheet.cell(row=latency_row, column=7)
            cell.value = "Latency Analysis"
            cell.font = Font(size=16, bold=True, color="00AA00")  # Green, bold, 12pt
            
            # Add the latency analysis text with word wrap
            cell = sheet.cell(row=latency_row, column=8)
            cell.value = latency_analysis
            cell.font = Font(size=14, color = "FF000080")
            cell.border = Border(left=Side(style='thin'),
                               right=Side(style='thin'),
                               top=Side(style='thin'),
                               bottom=Side(style='thin'))
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Calculate required row height for latency analysis
            latency_wrapped_lines = []
            for line in latency_analysis.split('\n'):
                latency_wrapped_lines.extend(textwrap.wrap(line, width=120))
            
            # Set row height for latency analysis section
            latency_row_height = max(25, len(latency_wrapped_lines) * 25)
            sheet.row_dimensions[latency_row].height = latency_row_height

        except Exception as e:
            print(f"Error adding analysis to summary sheet: {str(e)}")

        return sheet


    def create_graphs(self):
        if self.check_time_units():
            self.create_summary_sheet()
            self.create_multi_throughput_mb_graph()
            self.create_multi_throughput_records_graph()
            self.create_all_latency_compare_graphs()
            self.create_multi_latency_compare_graphs()
            self.create_multi_latency_graphs()
            self.create_multi_write_read_records_graph()
            self.create_multi_write_read_mb_graph()
            self.create_multi_write_read_timeout_events_graph()
            self.create_multi_write_read_timeout_events_per_sec_graph()
            self.create_total_multi_latency_percentile_graphs()
            self.create_total_multi_latency_percentile_count_graphs()
            self.create_total_mb_compare_graph()
            self.create_total_throughput_mb_compare_graph()
            self.create_total_throughput_records_compare_graph()
            self.create_total_min_latency_compare_graph()
            self.create_total_avg_latency_compare_graph()
            self.create_total_max_latency_compare_graph()
            self.create_total_write_read_timeout_events_compare_graph()
            self.wb.save(self.file)
            print("file : %s updated with graphs and AI documentation" % self.file)
