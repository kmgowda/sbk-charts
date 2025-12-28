#!/usr/local/bin/python3
# Copyright (c) KMG. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
##

"""sbk_charts.multicharts

Utilities to create charts that aggregate multiple SBK result sheets.

This module provides the SbkMultiCharts class which extends SbkCharts and
adds helpers to produce cross-sheet summary charts (multi-file or multi-run
comparisons). It reads multiple R and T sheets from the workbook, verifies
consistency (for example time units), and produces combined charts such as
aggregate latency comparisons, throughput summaries and histograms.

Notes
- This file only receives documentation additions; no functional code was
  changed.
"""

# sbk_charts :  Storage Benchmark Kit - Charts

from collections import OrderedDict
from datetime import date, datetime
from openpyxl.chart import Reference
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from ordered_set import OrderedSet

from src.sheets import constants
from .charts import SbkCharts


class SbkMultiCharts(SbkCharts):
    """Create multi-sheet charts by extending SbkCharts.

    This class adds routines that iterate over all R/T sheets in the workbook
    and produce aggregate charts that compare multiple runs or drivers.

    Typical usage
    - Instantiate with the SBK charts version and an existing workbook file.
    - Call `create_graphs()` to produce a full set of summary charts across
      all applicable sheets.

    Important behavior
    - `check_time_units()` will validate that all R-sheets use the same time
      unit; if mismatched, multi-sheet graphs will not be created.
    """

    def __init__(self, version, file):
        super().__init__(version, file)

    def check_time_units(self):
        """Validate that all R-sheet time units are identical.

        Returns
        - bool: True if all R-sheets share the same latency time unit, False
          otherwise.
        """
        ret = OrderedSet()
        for name in self.wb.sheetnames:
            if self.is_r_num_sheet(name):
                ret.add(self.get_time_unit(self.wb[name]))
        if len(ret) > 1:
            print("ERROR: Multiple Time unit are preset in " + self.file + " " + str(ret))
            return False
        print("Time Unit : " + ''.join(ret))
        return True

    def get_actions_storage_map(self):
        """Return a mapping of actions to the set of storage drivers encountered.

        Returns
        - OrderedDict[str, OrderedSet]: mapping where keys are action names and
          values are OrderedSet of storage names found in R-sheets.
        """
        ret = OrderedDict()
        for name in self.wb.sheetnames:
            if self.is_r_num_sheet(name):
                action = self.get_action_name(self.wb[name])
                if action not in ret:
                    ret[action] = OrderedSet()
                ret[action].add(self.get_storage_name(self.wb[name]))
        return ret

    def create_summary_sheet(self):
        """Create a human-friendly summary worksheet.

        The summary sheet contains metadata such as SBK Charts version, date
        and a short list of drivers/actions present in the workbook.

        Returns
        - openpyxl.worksheet.worksheet.Worksheet: the created summary sheet.
        """
        BLACK = 'FF000000'
        WHITE = 'FFFFFFFF'
        RED = 'FFFF0000'
        DARKRED = 'FF800000'
        BLUE = 'FF0000FF'
        DARKBLUE = 'FF000080'
        GREEN = 'FF00FF00'
        DARKGREEN = 'FF008000'
        YELLOW = 'FFFFFF00'
        DARKYELLOW = 'FF808000'

        acts = self.get_actions_storage_map()
        sheet = self.wb.create_sheet("Summary")
        row = 7
        col = 7
        sheet.column_dimensions[get_column_letter(col)].width = 45
        sheet.column_dimensions[get_column_letter(col + 1)].width = 75
        cell = sheet.cell(row, col + 1)
        cell.value = "SBK Charts "
        cell.font = Font(size="47", bold=True, color=DARKBLUE)
        cell.alignment = Alignment(horizontal='center')
        row += 1
        cell = sheet.cell(row, col)
        cell.value = "SBK Charts Version"
        cell.font = Font(size="27", bold=True, color=DARKRED)
        cell.alignment = Alignment(horizontal='left')
        cell = sheet.cell(row, col+1)
        cell.value = self.version
        cell.font = Font(size="27", bold=True, color=BLACK)
        cell.alignment = Alignment(horizontal='left')
        row += 1
        cell = sheet.cell(row, col)
        cell.value = "Date"
        cell.font = Font(size="18", bold=False, color=DARKRED)
        cell.alignment = Alignment(horizontal='left')
        cell = sheet.cell(row, col+1)
        cell.value =  date.today()
        cell.font = Font(size="18", bold=False, color=BLACK)
        cell.alignment = Alignment(horizontal='left')
        row +=1
        cell = sheet.cell(row, col)
        cell.value = "Time"
        cell.font = Font(size="18", bold=False, color=DARKRED)
        cell.alignment = Alignment(horizontal='left')
        cell = sheet.cell(row, col + 1)
        cell.value = datetime.now().strftime("%H:%M:%S")
        cell.font = Font(size="18", bold=False, color=BLACK)
        cell.alignment = Alignment(horizontal='left')
        row += 3
        drivers = OrderedSet()
        for values in acts.values():
            drivers.update(values)
        text = "Performance Analysis of Storage Drivers :  " + ", ".join(drivers)
        cell = sheet.cell(row, col)
        cell.value = text
        cell.font = Font(size="27", bold=True, color=RED)
        row += 1
        cell = sheet.cell(row, col)
        cell.value = "Time Unit"
        cell.font = Font(size="18", bold=False, color=BLUE)
        cell = sheet.cell(row, col + 1)
        cell.value = self.get_time_unit(self.wb[constants.R_PREFIX + "1"])
        cell.font = Font(size="18", bold=False, color=BLACK)
        row += 1
        for i, key in enumerate(acts):
            cell = sheet.cell(row + i, col)
            cell.value = key
            text = key
            cell.font = Font(size="18", bold=False, color=DARKGREEN)
            cell = sheet.cell(row + i, col + 1)
            cell.value = ", ".join(acts[key])
            cell.font = Font(size="18", bold=False, color=DARKRED)
            text += " : " + cell.value
            print(text)
        return sheet

    def create_all_latency_compare_graphs(self):
        """Generate combined latency comparison charts across R-sheets.

        Returns
        - list[Worksheet]: sheets that were created containing comparison charts.
        """
        charts, sheets = [], []
        for i in range(self.n_latency_charts):
            charts.append(self.create_latency_line_graph("Latency Variations"))
            sheets.append(self.wb.create_sheet("Latencies-" + str(i + 1)))
        for name in self.wb.sheetnames:
            if self.is_r_num_sheet(name):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                latency_series = self.get_latency_series(ws, prefix)
                for x in latency_series:
                    for i, g in enumerate(self.latency_groups):
                        if x in g:
                            charts[i].append(latency_series[x])
        for i, ch in enumerate(charts):
            ch.width = 70
            ch.height = 70
            sheets[i].add_chart(ch)
        return sheets

    def create_multi_latency_compare_graphs(self):
        """Create per-R-sheet latency comparison charts using the base class.

        This function delegates to `SbkCharts.create_latency_compare_graphs`
        for each R-sheet found in the workbook and aggregates the returned
        sheets into one list.

        Returns
        - list[Worksheet]
        """
        all_sheets = []
        for name in self.wb.sheetnames:
            if self.is_r_num_sheet(name):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                sheets = super().create_latency_compare_graphs(ws, prefix)
                all_sheets.extend(sheets)
        return all_sheets

    def create_multi_latency_graphs(self):
        """Create individual latency variation charts merged across R-sheets.

        Each latency metric gets one combined chart containing series from every
        R-sheet found in the workbook.

        Returns
        - list[Worksheet]
        """
        charts = OrderedDict()
        sheets = []
        for name in self.wb.sheetnames:
            if self.is_r_num_sheet(name):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                latency_series = self.get_latency_series(ws, prefix)
                for x in latency_series:
                    if x not in charts:
                        charts[x] = self.create_latency_line_graph(x + " Variations")
                    charts[x].append(latency_series[x])
        for x in charts:
            sheet = self.wb.create_sheet(x)
            sheet.add_chart(charts[x])
            sheets.append(sheet)
        return sheets

    def create_total_multi_latency_percentile_graphs(self):
        """Create combined percentile-line charts across T-sheets.

        For each percentile slice group (self.slc_percentile_names), this
        method constructs a LineChart and adds series from all T-sheets.

        Returns
        - list[Worksheet]
        """
        title = "Total Percentiles"
        sheets = []
        for i, names_list in enumerate(self.slc_percentile_names):
            chart = self.create_line_chart(title, "Percentiles", "Latency time in " + self.time_unit, 25, 50)
            x_labels = False
            for name in self.wb.sheetnames:
                if self.is_t_num_sheet(name):
                    ws = self.wb[name]
                    prefix = name + "_" + self.get_storage_name(ws)
                    latency_series = self.get_latency_percentile_series(ws, prefix, names_list)
                    for x in latency_series:
                        chart.append(latency_series[x])
                    if x_labels is False:
                        latency_cols = self.get_latency_percentile_columns(ws)
                        percentile_names = Reference(ws, min_col=latency_cols[names_list[0]], min_row=1,
                                                     max_col=latency_cols[names_list[-1]], max_row=1)
                        chart.set_categories(percentile_names)
                        x_labels = True
            sheet = self.wb.create_sheet("Total_Percentiles_" + str(i + 1))
            sheet.add_chart(chart)
            sheets.append(sheet)
        return sheets

    def create_total_multi_latency_percentile_count_graphs(self):
        """Create combined percentile-count histograms across T-sheets.

        Returns
        - Worksheet: the worksheet with the aggregated histogram chart.
        """
        title = "Total Percentiles Histogram"
        chart = self.create_bar_chart(title, "Percentiles", "Count", 25, 50)
        x_labels = False
        for name in self.wb.sheetnames:
            if self.is_t_num_sheet(name):
                ws = self.wb[name]
                prefix = name + "_" + self.get_storage_name(ws)
                latency_series = self.get_latency_percentile_count_series(ws, prefix, self.percentile_count_names)
                for x in latency_series:
                    chart.append(latency_series[x])
                if x_labels is False:
                    latency_cols = self.get_latency_percentile_count_columns(ws)
                    percentile_names = Reference(ws, min_col=latency_cols[self.percentile_count_names[0]], min_row=1,
                                                 max_col=latency_cols[self.percentile_count_names[-1]], max_row=1)
                    chart.set_categories(percentile_names)
                    x_labels = True
        sheet = self.wb.create_sheet("Total_Percentiles_Histogram" )
        sheet.add_chart(chart)


    def create_multi_throughput_mb_graph(self):
        """Create a combined MB/sec throughput variation chart from R-sheets.

        Returns
        - Worksheet with the combined MB/sec chart.
        """
        chart = self.create_line_chart("Throughput Variations in Mega Bytes / Seconds",
                                       "Intervals", "Throughput in MB/Sec", 25, 50)
        for name in self.wb.sheetnames:
            if self.is_r_num_sheet(name):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_throughput_write_request_mb_series(ws, prefix))
                chart.append(self.get_throughput_read_request_mb_series(ws, prefix))
                chart.append(self.get_throughput_mb_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("Throughput_MB")
        sheet.add_chart(chart)
        return sheet

    def create_multi_throughput_records_graph(self):
        """Create a combined records/sec throughput variation chart from R-sheets.

        Returns
        - Worksheet with the combined records/sec chart.
        """
        chart = self.create_line_chart("Throughput Variations in Records / Seconds",
                                       "Intervals", "Throughput in Records/Sec", 25, 50)
        for name in self.wb.sheetnames:
            if self.is_r_num_sheet(name):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_throughput_write_request_records_series(ws, prefix))
                chart.append(self.get_throughput_read_request_records_series(ws, prefix))
                chart.append(self.get_throughput_records_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("Throughput_Records")
        sheet.add_chart(chart)
        return sheet

    def create_multi_write_read_records_graph(self):
        """Create combined write/read record count charts across R-sheets.

        Returns
        - Worksheet with combined record-count variations.
        """
        chart = self.create_line_chart("Write and Read Records Variations",
                                       "Intervals", "Write and Read Records", 25, 50)
        for name in self.wb.sheetnames:
            if self.is_r_num_sheet(name):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_write_request_records_series(ws, prefix))
                chart.append(self.get_read_request_records_series(ws, prefix))
                chart.append(self.get_write_response_pending_records_series(ws, prefix))
                chart.append(self.get_read_response_pending_records_series(ws, prefix))
                chart.append(self.get_write_read_request_pending_records_series(ws, prefix))
                chart.append(self.get_records_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("Write_Read_Records")
        sheet.add_chart(chart)
        return sheet

    def create_multi_write_read_mb_graph(self):
        """Create combined write/read MB charts across R-sheets.

        Returns
        - Worksheet with combined MB variations.
        """
        chart = self.create_line_chart("Write and Read MBs Variations",
                                       "Intervals", "Write and Read MBs", 25, 50)
        for name in self.wb.sheetnames:
            if self.is_r_num_sheet(name):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_write_request_mb_series(ws, prefix))
                chart.append(self.get_read_request_mb_series(ws, prefix))
                chart.append(self.get_write_response_pending_mb_series(ws, prefix))
                chart.append(self.get_read_response_pending_mb_series(ws, prefix))
                chart.append(self.get_write_read_request_pending_mb_series(ws, prefix))
                chart.append(self.get_records_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("Write_Read_MB")
        sheet.add_chart(chart)
        return sheet

    def create_total_mb_compare_graph(self):
        """Create a bar chart that compares total MB values across T-sheets.

        Returns
        - Worksheet or None: the created sheet if data existed, otherwise None.
        """
        chart = None
        for name in self.wb.sheetnames:
            if self.is_t_num_sheet(name):
                ws = self.wb[name]
                if chart is None:
                    action = self.get_action_name(ws)
                    chart = self.create_bar_chart("Total Mega Bytes " + action, action, "MB", 25, 50)
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_write_request_mb_series(ws, prefix))
                chart.append(self.get_read_request_mb_series(ws, prefix))
                chart.append(self.get_write_response_pending_mb_series(ws, prefix))
                chart.append(self.get_read_response_pending_mb_series(ws, prefix))
                chart.append(self.get_write_read_request_pending_mb_series(ws, prefix))
                chart.append(self.get_mb_series(ws, prefix))
        if chart is not None:
            # add chart to the sheet
            sheet = self.wb.create_sheet("Total_MB")
            sheet.add_chart(chart)
            return sheet
        return None

    def create_multi_write_read_timeout_events_graph(self):
        """Create combined timeout events charts across R-sheets.

        Returns
        - Worksheet with combined timeout event variations.
        """
        chart = self.create_line_chart("Write and Read Timeout Events Variations",
                                       "Intervals", "Write and Read Timeout Events", 25, 50)
        for name in self.wb.sheetnames:
            if self.is_r_num_sheet(name):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_write_timeout_events_series(ws, prefix))
                chart.append(self.get_read_timeout_events_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("RW_TimeoutEvents")
        sheet.add_chart(chart)
        return sheet

    def create_multi_write_read_timeout_events_per_sec_graph(self):
        """Create combined timeout-events-per-second charts across R-sheets.

        Returns
        - Worksheet with combined timeout-events/sec variations.
        """
        chart = self.create_line_chart("Write and Read Timeout Events / Sec Variations",
                                       "Intervals", "Write and Read Timeout Events / Sec", 25, 50)
        for name in self.wb.sheetnames:
            if self.is_r_num_sheet(name):
                ws = self.wb[name]
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_write_timeout_events_per_sec_series(ws, prefix))
                chart.append(self.get_read_timeout_events_per_sec_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("RW_TimeoutEvents_Per_Sec")
        sheet.add_chart(chart)
        return sheet

    def create_total_throughput_mb_compare_graph(self):
        """Create a bar chart comparing total MB/sec throughput across T-sheets.

        Returns
        - Worksheet or None
        """
        chart = None
        for name in self.wb.sheetnames:
            if self.is_t_num_sheet(name):
                ws = self.wb[name]
                if chart is None:
                    action = self.get_action_name(ws)
                    chart = self.create_bar_chart("Total Throughput Variations in Mega Bytes / Seconds",
                                                  action, "Total Throughput in MB/Sec", 25, 50)
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_throughput_write_request_mb_series(ws, prefix))
                chart.append(self.get_throughput_read_request_mb_series(ws, prefix))
                chart.append(self.get_throughput_mb_series(ws, prefix))
        if chart is not None:
            # add chart to the sheet
            sheet = self.wb.create_sheet("Total_Throughput_MB")
            sheet.add_chart(chart)
            return sheet
        return None

    def create_total_throughput_records_compare_graph(self):
        """Create a bar chart comparing total records/sec throughput across T-sheets.

        Returns
        - Worksheet or None
        """
        chart = None
        for name in self.wb.sheetnames:
            if self.is_t_num_sheet(name):
                ws = self.wb[name]
                if chart is None:
                    action = self.get_action_name(ws)
                    chart = self.create_bar_chart("Total Throughput Variations in Records / Seconds",
                                                  action, "Total Throughput in Records/Sec", 25, 50)
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_throughput_write_request_records_series(ws, prefix))
                chart.append(self.get_throughput_read_request_records_series(ws, prefix))
                chart.append(self.get_throughput_records_series(ws, prefix))
        if chart is not None:
            # add chart to the sheet
            sheet = self.wb.create_sheet("Total_Throughput_Records")
            sheet.add_chart(chart)
            return sheet
        return None

    def create_total_avg_latency_compare_graph(self):
        """Create a bar chart comparing total average latency across T-sheets.

        Returns
        - Worksheet or None
        """
        chart = None
        for name in self.wb.sheetnames:
            if self.is_t_num_sheet(name):
                ws = self.wb[name]
                if chart is None:
                    action = self.get_action_name(ws)
                    chart = self.create_bar_chart("Total Average Latency Comparison",
                                                  action, "Total Average Latency", 25, 50)
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_avg_latency_series(ws, prefix))
        if chart is not None:
            # add chart to the sheet
            sheet = self.wb.create_sheet("Total_Avg_Latency")
            sheet.add_chart(chart)
            return sheet
        return None

    def create_total_min_latency_compare_graph(self):
        """Create a bar chart comparing total minimum latency across T-sheets.

        Returns
        - Worksheet or None
        """
        chart = None
        for name in self.wb.sheetnames:
            if self.is_t_num_sheet(name):
                ws = self.wb[name]
                if chart is None:
                    action = self.get_action_name(ws)
                    chart = self.create_bar_chart("Total Min Latency Comparison",
                                                  action, "Total Min Latency", 25, 50)
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_min_latency_series(ws, prefix))
        if chart is not None:
            # add chart to the sheet
            sheet = self.wb.create_sheet("Total_Min_Latency")
            sheet.add_chart(chart)
            return sheet
        return None

    def create_total_max_latency_compare_graph(self):
        """Create a bar chart comparing total maximum latency across T-sheets.

        Returns
        - Worksheet or None
        """
        chart = None
        for name in self.wb.sheetnames:
            if self.is_t_num_sheet(name):
                ws = self.wb[name]
                if chart is None:
                    action = self.get_action_name(ws)
                    chart = self.create_bar_chart("Total Max Latency Comparison",
                                                  action, "Total Max Latency", 25, 50)
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_max_latency_series(ws, prefix))
        if chart is not None:
            # add chart to the sheet
            sheet = self.wb.create_sheet("Total_Max_Latency")
            sheet.add_chart(chart)
            return sheet
        return None

    def create_total_write_read_timeout_events_compare_graph(self):
        """Create a bar chart comparing total write/read timeout events across T-sheets.

        Returns
        - Worksheet or None
        """
        chart = None
        for name in self.wb.sheetnames:
            if self.is_t_num_sheet(name):
                ws = self.wb[name]
                if chart is None:
                    action = self.get_action_name(ws)
                    chart = self.create_bar_chart("Total Write and Read Timeout Events Comparison",
                                                  action, "Write and Read Timeout Events", 25, 50)
                prefix = name + "-" + self.get_storage_name(ws)
                chart.append(self.get_write_timeout_events_series(ws, prefix))
                chart.append(self.get_read_timeout_events_series(ws, prefix))
        if chart is not None:
            # add chart to the sheet
            sheet = self.wb.create_sheet("Total_RW_TimeoutEvents")
            sheet.add_chart(chart)
            return sheet
        return None


    def create_graphs(self):
        """Top-level multi-sheet graph generation.

        This method will run a sequence of multi-sheet chart creation routines
        after verifying consistency (time units). If checks pass the modified
        workbook will be saved.
        """
        if self.check_time_units():
            self.create_multi_throughput_records_graph()
            self.create_all_latency_compare_graphs()
            self.create_multi_latency_compare_graphs()
            self.create_multi_latency_graphs()
            self.create_multi_write_read_records_graph()
            self.create_multi_write_read_mb_graph()
            self.create_multi_write_read_timeout_events_graph()
            self.create_multi_write_read_timeout_events_per_sec_graph()
            self.create_total_multi_latency_percentile_graphs()
            self.create_total_multi_latency_percentile_count_graphs()
            self.create_total_mb_compare_graph()
            self.create_total_throughput_mb_compare_graph()
            self.create_total_throughput_records_compare_graph()
            self.create_total_min_latency_compare_graph()
            self.create_total_avg_latency_compare_graph()
            self.create_total_max_latency_compare_graph()
            self.create_total_write_read_timeout_events_compare_graph()
            self.wb.save(self.file)
            print("file : %s updated with graphs" % self.file)
