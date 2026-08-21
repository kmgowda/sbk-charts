#!/usr/local/bin/python3
# Copyright (c) KMG. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
##

"""SBK Charts Module

This module provides the SbkCharts class for generating visual representations
of storage benchmark results in Excel workbooks. It creates various charts
including latency percentiles, throughput metrics (MB/sec and records/sec),
and percentile distribution histograms.

Key Features:
- Line and bar chart generation for performance metrics
- Support for multiple storage systems and test scenarios
- Customizable chart styling and formatting
- Integration with openpyxl for Excel workbook manipulation

Note: This module focuses on visualization only and does not modify the
underlying benchmark data.
"""

# Storage Benchmark Kit - Charts Module
import os
from collections import OrderedDict
from copy import copy
from typing import final

from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.chart.text import RichText
from openpyxl.drawing.image import Image
from openpyxl.drawing.text import (
    CharacterProperties,
    Paragraph,
    ParagraphProperties,
    RichTextProperties,
)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl_image_loader import SheetImageLoader

import src.sheets.constants as sheets_constants
from src.charts import constants
from src.charts.utils import (
    get_columns_from_worksheet,
    get_storage_name_from_worksheet,
    get_time_unit_from_worksheet,
    is_r_num_sheet,
    is_t_num_sheet,
)
from src.version.sbk_version import __sbk_version__


# Excel chart dimensions are expressed in centimetres by openpyxl. These
# dimensions produce a large, readable 16:9 chart on a dedicated worksheet.
DEFAULT_CHART_HEIGHT = 27
DEFAULT_CHART_WIDTH = 48

CHART_TITLE_FONT_SIZE = 3400
AXIS_TITLE_FONT_SIZE = 2200
AXIS_LABEL_FONT_SIZE = 1800
LEGEND_FONT_SIZE = 1600
CHART_LINE_WIDTH = 44450
CHART_MARKER_SIZE = 8
CHART_TEXT_COLOR = "374151"

INTERVALS_AXIS_TITLE = "Intervals"
PERCENTILES_AXIS_TITLE = "Percentiles"
COUNT_AXIS_TITLE = "Count"
LATENCY_AXIS_TITLE_PREFIX = "Latency time in "

TABLE_BODY_FONT_SIZE = 15
TABLE_HEADER_FONT_SIZE = 19
TABLE_HEADER_FILLS = (
    "D9EAFE",
    "DAF3F3",
    "EDE2F4",
    "FFE4D1",
    "E2F1E2",
    "E3E6FF",
)
TABLE_HEADER_TEXT_COLORS = (
    "12355B",
    "0F4C5C",
    "4A235A",
    "7A3512",
    "234D2C",
    "283593",
)
TABLE_COLUMN_FILLS = (
    "EAF2FF",
    "E8F7F7",
    "F4ECFA",
    "FFF1E6",
    "EDF7ED",
    "EEF0FF",
)
TABLE_GRID_COLOR = "C8D6E8"
TABLE_MIN_COLUMN_WIDTH = 10
TABLE_MAX_COLUMN_WIDTH = 60
TABLE_COLUMN_SCALE = 1.25
TABLE_HEADER_WIDTH_SCALE = 1.45
TABLE_HEADER_WIDTH_PADDING = 4

# Saturated, high-contrast colours make adjacent series easy to distinguish.
# Dash patterns and marker shapes add non-colour cues for accessibility.
CHART_SERIES_COLORS = (
    "0066CC",
    "FF7A00",
    "009E60",
    "D81B60",
    "6A1B9A",
    "00A6D6",
    "C62828",
    "7CB342",
    "8D5A00",
    "3949AB",
)
CHART_DASH_STYLES = ("solid", "dash", "sysDot", "dashDot")
CHART_MARKER_SYMBOLS = (
    "circle",
    "square",
    "diamond",
    "triangle",
    "x",
    "plus",
    "star",
    "dot",
    "dash",
)


class SbkCharts:
    """Create and manage Excel charts for SBK results.

    Responsibilities
    - Load an existing workbook and inspect R/T sheets for SBK results.
    - Provide helper methods to build Series objects from result columns.
    - Create and style charts (titles, axis labels, sizes) and insert them
      into new worksheets in the workbook.

    Initialization
    - file: path to the workbook file to open and modify

    Public highlights
    - create_graphs(): top-level routine to produce the common set of charts
      (latency, throughput, percentiles) and save the workbook.

    Note: This docstring only documents behavior; no code paths are altered.
    """

    def __init__(self,  file):
        """Load workbook and initialize derived parameters.

        Parameters
        - file (str): path to the Excel workbook file to open

        Returns
        - None

        Raises
        - FileNotFoundError: if the specified file does not exist
        - InvalidFileException: if the file is not a valid Excel workbook
        """
        self.version = __sbk_version__
        self.file = file
        self.wb = load_workbook(self.file)
        self.time_unit = get_time_unit_from_worksheet(
            self.wb[sheets_constants.FIRST_RESULT_SHEET]
        )
        self.latency_groups = constants.LATENCY_GROUPS
        self.n_latency_charts = len(self.latency_groups)
        self.slc_percentile_names = constants.SLC_PERCENTILE_GROUPS
        self.percentile_count_names = constants.PERCENTILE_COUNT_COLUMNS

    def get_latency_percentile_columns(self, ws):
        """Return only the latency percentile columns (names starting with
        'Percentile_' but not 'Percentile_Count').

        Parameters
        - ws: openpyxl worksheet

        Returns
        - OrderedDict[str, int]: mapping of percentile column name to index
        """
        columns = get_columns_from_worksheet(ws)
        ret = OrderedDict()
        for key in columns.keys():
            if key.startswith(constants.PERCENTILE_PREFIX) and not key.startswith(
                constants.PERCENTILE_COUNT_PREFIX
            ):
                ret[key] = columns[key]
        return ret

    def get_latency_percentile_count_columns(self, ws):
        """Return only the percentile count columns (names starting with
        'Percentile_Count').

        Parameters
        - ws: openpyxl worksheet

        Returns
        - OrderedDict[str, int]: mapping of percentile count column name to index
        """
        columns = get_columns_from_worksheet(ws)
        ret = OrderedDict()
        for key in columns.keys():
            if key.startswith(constants.PERCENTILE_COUNT_PREFIX):
                ret[key] = columns[key]
        return ret

    def get_latency_columns(self, ws):
        """Return an OrderedDict of the main latency-related columns: AVG, MIN,
        MAX and percentile columns.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - OrderedDict[str, int]
        """
        columns = get_columns_from_worksheet(ws)
        ret = OrderedDict()
        ret[constants.AVG_LATENCY] = columns[constants.AVG_LATENCY]
        ret[constants.MIN_LATENCY] = columns[constants.MIN_LATENCY]
        ret[constants.MAX_LATENCY] = columns[constants.MAX_LATENCY]
        ret.update(self.get_latency_percentile_columns(ws))
        return ret

    @final
    def __add_chart_attributes(self, chart, title, x_title, y_title, height, width):
        """Apply common visual attributes to a chart object.

        This sets the chart and axis titles (with font sizes and boldness),
        chart dimensions, and makes sure axes are visible.

        Parameters
        - chart: an openpyxl chart object (LineChart or BarChart)
        - title (str): chart title text
        - x_title (str): label for x axis
        - y_title (str): label for y axis
        - height (int): chart height (openpyxl units)
        - width (int): chart width (openpyxl units)

        Returns
        - None (modifies chart object in-place)
        """
        # Use one clean built-in theme as a base. Explicit font and series
        # settings below keep the charts readable in Excel-compatible viewers
        # that only partially support built-in chart styles.
        chart.style = 10
        chart.roundedCorners = False
        chart.display_blanks = "gap"

        chart.title = title
        chart.title.tx.rich.p[0].pPr.defRPr = CharacterProperties(
            sz=CHART_TITLE_FONT_SIZE,
            b=True,
            solidFill="1F2937",
        )
        
        chart.x_axis.title = x_title
        if not hasattr(chart.x_axis.title.tx.rich.p[0], 'pPr'):
            chart.x_axis.title.tx.rich.p[0].pPr = ParagraphProperties()
        chart.x_axis.title.tx.rich.p[0].pPr.defRPr = CharacterProperties(
            sz=AXIS_TITLE_FONT_SIZE,
            b=True,
            solidFill=CHART_TEXT_COLOR,
        )
        
        chart.y_axis.title = y_title
        if not hasattr(chart.y_axis.title.tx.rich.p[0], 'pPr'):
            chart.y_axis.title.tx.rich.p[0].pPr = ParagraphProperties()
        chart.y_axis.title.tx.rich.p[0].pPr.defRPr = CharacterProperties(
            sz=AXIS_TITLE_FONT_SIZE,
            b=True,
            solidFill=CHART_TEXT_COLOR,
        )

        chart.height = height
        chart.width = width
        chart.anchor = "B2"
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.x_axis.tickLblPos = "low"
        chart.y_axis.tickLblPos = "nextTo"
        chart.x_axis.majorTickMark = "out"
        chart.y_axis.majorTickMark = "out"
        chart.x_axis.txPr = self.__text_properties(AXIS_LABEL_FONT_SIZE)
        chart.y_axis.txPr = self.__text_properties(AXIS_LABEL_FONT_SIZE)

        if chart.legend is not None:
            chart.legend.position = "b"
            chart.legend.overlay = False
            chart.legend.txPr = self.__text_properties(LEGEND_FONT_SIZE)

    @staticmethod
    def __text_properties(font_size):
        """Create reusable DrawingML text properties for chart labels."""
        character_properties = CharacterProperties(
            lang="en-US",
            sz=font_size,
            solidFill=CHART_TEXT_COLOR,
        )
        return RichText(
            bodyPr=RichTextProperties(),
            p=[
                Paragraph(
                    pPr=ParagraphProperties(defRPr=character_properties),
                    endParaRPr=character_properties,
                )
            ],
        )

    @final
    def apply_chart_theme(self):
        """Finalize chart layout and series styling across the workbook.

        Chart factories run before data series are attached. This final pass
        can therefore adapt legend placement and line/marker styling to the
        actual number of series in every chart.
        """
        for sheet in self.wb.worksheets:
            if sheet._charts:
                sheet.sheet_view.showGridLines = False
                sheet.sheet_view.zoomScale = 80
            for chart in sheet._charts:
                series_count = len(chart.series)
                if chart.legend is not None:
                    chart.legend.position = "r" if series_count > 6 else "b"
                    chart.legend.overlay = False
                    chart.legend.txPr = self.__text_properties(LEGEND_FONT_SIZE)

                for index, series in enumerate(chart.series):
                    color = CHART_SERIES_COLORS[index % len(CHART_SERIES_COLORS)]
                    series.graphicalProperties.line.solidFill = color

                    if isinstance(chart, LineChart):
                        series.graphicalProperties.line.width = CHART_LINE_WIDTH
                        palette_cycle = index // len(CHART_SERIES_COLORS)
                        series.graphicalProperties.line.prstDash = (
                            CHART_DASH_STYLES[
                                palette_cycle % len(CHART_DASH_STYLES)
                            ]
                        )
                        if series_count <= len(CHART_SERIES_COLORS):
                            series.marker.symbol = CHART_MARKER_SYMBOLS[
                                index % len(CHART_MARKER_SYMBOLS)
                            ]
                            series.marker.size = CHART_MARKER_SIZE
                            series.marker.graphicalProperties.solidFill = color
                            series.marker.graphicalProperties.line.solidFill = color
                    elif isinstance(chart, BarChart):
                        series.graphicalProperties.solidFill = color

                if isinstance(chart, BarChart):
                    chart.type = "col"
                    chart.grouping = "clustered"
                    chart.overlap = 0
                    chart.gapWidth = 55
                    chart.varyColors = False

    @final
    def apply_table_theme(self):
        """Increase workbook table readability and apply consistent styling."""
        thin_bottom = Side(style="thin", color=TABLE_GRID_COLOR)

        for sheet in self.wb.worksheets:
            # Preserve intentionally large/custom Summary typography while
            # ensuring every populated workbook cell has a readable minimum.
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    current_size = cell.font.sz or 11
                    if current_size < TABLE_BODY_FONT_SIZE:
                        font = copy(cell.font)
                        font.sz = TABLE_BODY_FONT_SIZE
                        cell.font = font

            is_table = (
                is_r_num_sheet(sheet.title)
                or is_t_num_sheet(sheet.title)
                or sheet.title == sheets_constants.DURATIONS_SHEET
            )
            if not is_table:
                continue

            sheet.freeze_panes = "A2"
            # Keep table headers clean: filtering dropdowns are intentionally
            # disabled in generated reports.
            sheet.auto_filter.ref = None
            sheet.sheet_view.zoomScale = 85

            for column_index in range(1, sheet.max_column + 1):
                column_letter = get_column_letter(column_index)
                current_width = sheet.column_dimensions[column_letter].width or 13
                header_value = sheet.cell(row=1, column=column_index).value
                header_width = (
                    len(str(header_value)) * TABLE_HEADER_WIDTH_SCALE
                    + TABLE_HEADER_WIDTH_PADDING
                    if header_value is not None
                    else TABLE_MIN_COLUMN_WIDTH
                )
                sheet.column_dimensions[column_letter].width = min(
                    max(
                        current_width * TABLE_COLUMN_SCALE,
                        header_width,
                        TABLE_MIN_COLUMN_WIDTH,
                    ),
                    TABLE_MAX_COLUMN_WIDTH,
                )

            sheet.row_dimensions[1].height = 36
            for cell in sheet[1]:
                header_fill = TABLE_HEADER_FILLS[
                    (cell.column - 1) % len(TABLE_HEADER_FILLS)
                ]
                header_text_color = TABLE_HEADER_TEXT_COLORS[
                    (cell.column - 1) % len(TABLE_HEADER_TEXT_COLORS)
                ]
                cell.font = Font(
                    name=cell.font.name or "Calibri",
                    size=TABLE_HEADER_FONT_SIZE,
                    bold=True,
                    color=header_text_color,
                )
                cell.fill = PatternFill(fill_type="solid", fgColor=header_fill)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=False,
                )
                cell.border = Border(bottom=thin_bottom)

            for row_index in range(2, sheet.max_row + 1):
                sheet.row_dimensions[row_index].height = 26
                for cell in sheet[row_index]:
                    if cell.value is None:
                        continue
                    column_fill = TABLE_COLUMN_FILLS[
                        (cell.column - 1) % len(TABLE_COLUMN_FILLS)
                    ]
                    font = copy(cell.font)
                    font.sz = max(font.sz or 11, TABLE_BODY_FONT_SIZE)
                    font.color = "1F2937"
                    cell.font = font
                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor=column_fill,
                    )
                    alignment = copy(cell.alignment)
                    alignment.vertical = "center"
                    cell.alignment = alignment
                    cell.border = Border(bottom=thin_bottom)

    @final
    def create_line_chart(
        self,
        title,
        x_title,
        y_title,
        height=DEFAULT_CHART_HEIGHT,
        width=DEFAULT_CHART_WIDTH,
    ):
        """Factory for LineChart with common attributes applied.

        Parameters
        - title (str): chart title
        - x_title (str): x axis title
        - y_title (str): y axis title
        - height (int): height units
        - width (int): width units

        Returns
        - openpyxl.chart.LineChart: configured LineChart object
        """
        chart = LineChart()
        self.__add_chart_attributes(chart, title, x_title, y_title, height, width)
        return chart

    @final
    def create_bar_chart(
        self,
        title,
        x_title,
        y_title,
        height=DEFAULT_CHART_HEIGHT,
        width=DEFAULT_CHART_WIDTH,
    ):
        """Factory for BarChart with common attributes applied.

        Returns a BarChart object with titles and dimensions set.
        """
        chart = BarChart()
        self.__add_chart_attributes(chart, title, x_title, y_title, height, width)
        return chart

    def create_latency_line_graph(self, title):
        """Convenience wrapper to create a latency-specific LineChart.

        The y-axis title will include the currently configured time unit.
        """
        return self.create_line_chart(
            title,
            INTERVALS_AXIS_TITLE,
            LATENCY_AXIS_TITLE_PREFIX + self.time_unit,
        )

    def get_latency_series(self, ws, ws_name):
        """Build Series objects for all latency columns in the worksheet.

        Parameters
        - ws: openpyxl worksheet containing latency columns
        - ws_name (str): prefix used in series titles (usually sheet name)

        Returns
        - OrderedDict[str, Series]: mapping of column name to Series instance

        Notes
        - Series titles are formed as '<ws_name>-<column_name>' to make them
          identifiable when charts display legends.
        """
        latencies = self.get_latency_columns(ws)
        data_series = OrderedDict()
        for x in latencies:
            data_series[x] = Series(Reference(ws, min_col=latencies[x], min_row=2,
                                              max_col=latencies[x], max_row=ws.max_row),
                                    title=ws_name + "-" + x)
        return data_series

    def get_latency_percentile_series(self, ws, ws_name, names_list):
        """Create row-wise Series that contain percentile values across columns.

        Each returned Series corresponds to one row (one measurement interval)
        and contains percentile values from min_col..max_col as a single series.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix
        - names_list (list[str]): ordered list of percentile column names

        Returns
        - OrderedDict[int, Series]: mapping of row number to Series object
        """
        latencies = self.get_latency_percentile_columns(ws)
        data_series = OrderedDict()
        min_col = latencies[names_list[0]]
        max_col = latencies[names_list[-1]]
        for r in range(2, ws.max_row + 1):
            data_series[r] = Series(Reference(ws, min_col=min_col, min_row=r, max_col=max_col, max_row=r),
                                    title=ws_name + "_" + str(r-1))
        return data_series

    def get_latency_percentile_count_series(self, ws, ws_name, names_list):
        """Create row-wise Series for percentile count histogram generation.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix
        - names_list (list[str]): ordered list of percentile-count column names

        Returns
        - OrderedDict[int, Series]
        """
        latencies = self.get_latency_percentile_count_columns(ws)
        data_series = OrderedDict()
        min_col = latencies[names_list[0]]
        max_col = latencies[names_list[-1]]
        for r in range(2, ws.max_row + 1):
            data_series[r] = Series(Reference(ws, min_col=min_col, min_row=r, max_col=max_col, max_row=r),
                                    title=ws_name + "_" + str(r-1))
        return data_series

    @final
    def __get_column_values(self, ws, column_name):
        """Utility to read a full column (skipping header) into a Python list.

        Parameters
        - ws: openpyxl worksheet
        - column_name (str): header name to look up

        Returns
        - list: cell values from row 2 to ws.max_row inclusive
        """
        cols = get_columns_from_worksheet(ws)
        values = []
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=cols[column_name]).value
            values.append(cell_value)
        return values

    @final
    def get_throughput_mb_values(self, ws):
        """Return the list of throughput (MB/sec) values from the worksheet.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]: list of throughput values in MB/sec
        """
        return self.__get_column_values(ws, constants.MB_PER_SEC)

    @final
    def get_throughput_write_request_mb_values(self, ws):
        """Return the list of write request throughput (MB/sec) values.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.WRITE_REQUEST_MB_PER_SEC)

    @final
    def get_throughput_read_request_mb_values(self, ws):
        """Return the list of read request throughput (MB/sec) values.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.READ_REQUEST_MB_PER_SEC)

    @final
    def get_throughput_records_values(self, ws):
        """Return the list of throughput (records/sec) values from the worksheet.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]: list of throughput values in records/sec
        """
        return self.__get_column_values(ws, constants.RECORDS_PER_SEC)

    @final
    def get_throughput_write_request_records_values(self, ws):
        """Return the list of write request throughput (records/sec) values.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.WRITE_REQUEST_RECORDS_PER_SEC)

    @final
    def get_throughput_read_request_records_values(self, ws):
        """Return the list of read request throughput (records/sec) values.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.READ_REQUEST_RECORDS_PER_SEC)

    @final
    def get_records_values(self, ws):
        """Return the list of record count values from the worksheet.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[int]: list of record counts
        """
        return self.__get_column_values(ws, constants.RECORDS)

    @final
    def get_mb_values(self, ws):
        """Return the list of MB values from the worksheet.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]: list of MB values
        """
        return self.__get_column_values(ws, constants.MB)

    @final
    def get_write_request_mb_values(self, ws):
        """Return the list of write request MB values.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.WRITE_REQUEST_MB)

    @final
    def get_write_request_records_values(self, ws):
        """Return the list of write request record values.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.WRITE_REQUEST_RECORDS)

    @final
    def get_write_response_pending_mb_values(self, ws):
        """Return the list of pending write response MB values.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.WRITE_RESPONSE_PENDING_MB)

    @final
    def get_write_response_pending_records_values(self, ws):
        """Return the list of pending write response record values.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.WRITE_RESPONSE_PENDING_RECORDS)

    @final
    def get_read_request_records_values(self, ws):
        """Return the list of read request record values.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.READ_REQUEST_RECORDS)

    @final
    def get_read_request_mb_values(self, ws):
        """Return the list of read request MB values.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.READ_REQUEST_MB)

    @final
    def get_read_response_pending_mb_values(self, ws):
        """Return the list of pending read response MB values.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.READ_RESPONSE_PENDING_MB)

    @final
    def get_read_response_pending_records_values(self, ws):
        """Return the list of pending read response record values.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.READ_RESPONSE_PENDING_RECORDS)

    @final
    def get_write_read_request_pending_mb_values(self, ws):
        """Return the list of pending write/read request MB values.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.WRITE_READ_REQUEST_PENDING_MB)

    @final
    def get_write_read_request_pending_records_values(self, ws):
        """Return the list of pending write/read request record values.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.WRITE_READ_REQUEST_PENDING_RECORDS)

    @final
    def get_avg_latency_values(self, ws, ws_name):
        """Return the list of average latency values.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): worksheet name (for series title)

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.AVG_LATENCY)

    @final
    def get_min_latency_values(self, ws, ws_name):
        """Return the list of minimum latency values.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): worksheet name (for series title)

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.MIN_LATENCY)

    @final
    def get_max_latency_values(self, ws, ws_name):
        """Return the list of maximum latency values.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): worksheet name (for series title)

        Returns
        - list[float]
        """
        return self.__get_column_values(ws, constants.MAX_LATENCY)

    @final
    def get_write_timeout_events_values(self, ws):
        """Return the list of write timeout event counts.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[int]: list of write timeout event counts
        """
        return self.__get_column_values(ws, constants.WRITE_TIMEOUT_EVENTS)

    @final
    def get_write_timeout_events_per_sec_values(self, ws):
        """Return the list of write timeout events per second values.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]: list of write timeout events per second
        """
        return self.__get_column_values(ws, constants.WRITE_TIMEOUT_EVENTS_PER_SEC)

    @final
    def get_read_timeout_events_values(self, ws):
        """Return the list of read timeout event counts.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[int]: list of read timeout event counts
        """
        return self.__get_column_values(ws, constants.READ_TIMEOUT_EVENTS)

    @final
    def get_read_timeout_events_per_sec_values(self, ws):
        """Return the list of read timeout events per second values.

        Parameters
        - ws: openpyxl worksheet

        Returns
        - list[float]: list of read timeout events per second
        """
        return self.__get_column_values(ws, constants.READ_TIMEOUT_EVENTS_PER_SEC)

    @final
    def __get_column_series(self, ws, ws_name, column_name):
        """Create a Series object for a single column of data.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix
        - column_name (str): name of the column to reference

        Returns
        - Series: openpyxl Series object for the column

        Notes
        - Series title is formed as '<ws_name>-<column_name>'.
        """
        cols = get_columns_from_worksheet(ws)
        return Series(Reference(ws, min_col=cols[column_name], min_row=2,
                                max_col=cols[column_name], max_row=ws.max_row),
                      title=ws_name + "-" + column_name)

    @final
    def get_throughput_mb_series(self, ws, ws_name):
        """Return a Series for the throughput (MB/sec) column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.MB_PER_SEC)

    @final
    def get_throughput_write_request_mb_series(self, ws, ws_name):
        """Return a Series for the write request throughput (MB/sec) column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.WRITE_REQUEST_MB_PER_SEC)

    @final
    def get_throughput_read_request_mb_series(self, ws, ws_name):
        """Return a Series for the read request throughput (MB/sec) column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.READ_REQUEST_MB_PER_SEC)

    @final
    def get_throughput_records_series(self, ws, ws_name):
        """Return a Series for the throughput (records/sec) column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.RECORDS_PER_SEC)

    @final
    def get_throughput_write_request_records_series(self, ws, ws_name):
        """Return a Series for the write request throughput (records/sec) column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.WRITE_REQUEST_RECORDS_PER_SEC)

    @final
    def get_throughput_read_request_records_series(self, ws, ws_name):
        """Return a Series for the read request throughput (records/sec) column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.READ_REQUEST_RECORDS_PER_SEC)

    @final
    def get_records_series(self, ws, ws_name):
        """Return a Series for the record count column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.RECORDS)

    @final
    def get_mb_series(self, ws, ws_name):
        """Return a Series for the MB column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.MB)

    @final
    def get_write_request_mb_series(self, ws, ws_name):
        """Return a Series for the write request MB column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.WRITE_REQUEST_MB)

    @final
    def get_write_request_records_series(self, ws, ws_name):
        """Return a Series for the write request record column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.WRITE_REQUEST_RECORDS)

    @final
    def get_write_response_pending_mb_series(self, ws, ws_name):
        """Return a Series for the pending write response MB column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.WRITE_RESPONSE_PENDING_MB)

    @final
    def get_write_response_pending_records_series(self, ws, ws_name):
        """Return a Series for the pending write response record column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.WRITE_RESPONSE_PENDING_RECORDS)

    @final
    def get_read_request_records_series(self, ws, ws_name):
        """Return a Series for the read request record column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.READ_REQUEST_RECORDS)

    @final
    def get_read_request_mb_series(self, ws, ws_name):
        """Return a Series for the read request MB column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.READ_REQUEST_MB)

    @final
    def get_read_response_pending_mb_series(self, ws, ws_name):
        """Return a Series for the pending read response MB column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.READ_RESPONSE_PENDING_MB)

    @final
    def get_read_response_pending_records_series(self, ws, ws_name):
        """Return a Series for the pending read response record column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.READ_RESPONSE_PENDING_RECORDS)

    @final
    def get_write_read_request_pending_mb_series(self, ws, ws_name):
        """Return a Series for the pending write/read request MB column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.WRITE_READ_REQUEST_PENDING_MB)

    @final
    def get_write_read_request_pending_records_series(self, ws, ws_name):
        """Return a Series for the pending write/read request record column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.WRITE_READ_REQUEST_PENDING_RECORDS)

    @final
    def get_avg_latency_series(self, ws, ws_name):
        """Return a Series for the average latency column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.AVG_LATENCY)

    @final
    def get_min_latency_series(self, ws, ws_name):
        """Return a Series for the minimum latency column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.MIN_LATENCY)

    @final
    def get_max_latency_series(self, ws, ws_name):
        """Return a Series for the maximum latency column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.MAX_LATENCY)

    @final
    def get_write_timeout_events_series(self, ws, ws_name):
        """Return a Series for the write timeout events column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.WRITE_TIMEOUT_EVENTS)

    @final
    def get_write_timeout_events_per_sec_series(self, ws, ws_name):
        """Return a Series for the write timeout events per second column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.WRITE_TIMEOUT_EVENTS_PER_SEC)

    @final
    def get_read_timeout_events_series(self, ws, ws_name):
        """Return a Series for the read timeout events column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.READ_TIMEOUT_EVENTS)

    @final
    def get_read_timeout_events_per_sec_series(self, ws, ws_name):
        """Return a Series for the read timeout events per second column.

        Parameters
        - ws: openpyxl worksheet
        - ws_name (str): series title prefix

        Returns
        - Series
        """
        return self.__get_column_series(ws, ws_name, constants.READ_TIMEOUT_EVENTS_PER_SEC)

    def create_latency_compare_graphs(self, ws, prefix):
        """Create and return a set of latency comparison graphs (line charts).

        This produces a set of line charts comparing different latency percentiles
        across the intervals measured. Each chart group is placed in a new worksheet.

        Parameters
        - ws: openpyxl worksheet containing the source data
        - prefix (str): prefix for naming the new sheets

        Returns
        - list[Worksheet]: list of newly created worksheets with charts

        Notes
        - Exactly `self.n_latency_charts` sheets/charts are created.
        """
        charts, sheets = [], []
        for i in range(self.n_latency_charts):
            charts.append(self.create_latency_line_graph("Latency Variations"))
            sheets.append(self.wb.create_sheet("Latencies-" + prefix + "-" + str(i + 1)))
        latency_series = self.get_latency_series(ws, prefix)
        for x in latency_series:
            for i, g in enumerate(self.latency_groups):
                if x in g:
                    charts[i].append(latency_series[x])
        for i, ch in enumerate(charts):
            sheets[i].add_chart(ch)
        return sheets

    def create_latency_graphs(self, ws, prefix):
        """Create individual latency variation graphs for each latency metric.

        Parameters
        - ws: openpyxl worksheet containing the source data
        - prefix (str): prefix for naming the new sheets

        Returns
        - list[Worksheet]: list of newly created worksheets with charts
        """
        sheets = []
        latency_series = self.get_latency_series(ws, prefix)
        for x in latency_series:
            chart = self.create_latency_line_graph(x + " Variations")
            # adding data
            chart.append(latency_series[x])
            # add chart to the sheet
            sheet = self.wb.create_sheet(x)
            sheet.add_chart(chart)
            sheets.append(sheet)
        return sheets

    def create_total_latency_percentile_graphs(self, ws, prefix):
        """Create graphs for total latency percentiles across all intervals.

        Parameters
        - ws: openpyxl worksheet containing the source data
        - prefix (str): prefix for naming the new sheets

        Returns
        - list[Worksheet]: list of newly created worksheets with charts
        """
        title = "Total Percentiles"
        latency_cols = self.get_latency_percentile_columns(ws)
        sheets = []
        for i, percentile_names in enumerate(self.slc_percentile_names):
            chart = self.create_line_chart(
                title,
                PERCENTILES_AXIS_TITLE,
                LATENCY_AXIS_TITLE_PREFIX + self.time_unit,
            )
            latency_series = self.get_latency_percentile_series(ws, prefix, percentile_names)
            for x in latency_series:
                chart.append(latency_series[x])
            # Add x-axis labels
            percentiles = Reference(ws, min_col=latency_cols[percentile_names[0]], min_row=1,
                                    max_col=latency_cols[percentile_names[-1]], max_row=1)
            chart.set_categories(percentiles)
            # add chart to the sheet
            sheet = self.wb.create_sheet("Total_Percentiles_" + str(i + 1))
            sheet.add_chart(chart)
            sheets.append(sheet)
        return sheets

    def create_total_latency_percentile_count_graphs(self, ws, prefix):
        """Create a histogram graph for total latency percentile counts.

        Parameters
        - ws: openpyxl worksheet containing the source data
        - prefix (str): prefix for naming the new sheet

        Returns
        - Worksheet: the newly created worksheet with the chart
        """
        title = "Total Percentiles Histogram"
        latency_cols = self.get_latency_percentile_count_columns(ws)
        chart = self.create_bar_chart(
            title,
            PERCENTILES_AXIS_TITLE,
            COUNT_AXIS_TITLE,
        )
        latency_series = self.get_latency_percentile_count_series(ws, prefix, self.percentile_count_names)
        for x in latency_series:
            chart.append(latency_series[x])
        # Add x-axis labels
        percentiles_counts = Reference(ws, min_col=latency_cols[self.percentile_count_names[0]], min_row=1,
                                max_col=latency_cols[self.percentile_count_names[-1]], max_row=1)
        chart.set_categories(percentiles_counts)
        # add chart to the sheet
        sheet = self.wb.create_sheet("Total_Percentiles_Histogram" )
        sheet.add_chart(chart)
        return sheet


    def create_throughput_mb_graph(self, ws, prefix):
        """Create a throughput (MB/sec) variation graph.

        This graph shows how the throughput in MB/sec varies across the
        different intervals measured.

        Parameters
        - ws: openpyxl worksheet containing the source data
        - prefix (str): prefix for naming the new sheet

        Returns
        - Worksheet: the newly created worksheet with the chart
        """
        chart = self.create_line_chart(
            "Throughput Variations in Mega Bytes / Seconds",
            INTERVALS_AXIS_TITLE,
            "Throughput in MB/Sec",
        )
        # adding data
        chart.append(self.get_throughput_write_request_mb_series(ws, prefix))
        chart.append(self.get_throughput_read_request_mb_series(ws, prefix))
        chart.append(self.get_throughput_mb_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("MB_Sec")
        sheet.add_chart(chart)
        return sheet

    def create_throughput_records_graph(self, ws, prefix):
        """Create a throughput (records/sec) variation graph.

        This graph shows how the throughput in records/sec varies across the
        different intervals measured.

        Parameters
        - ws: openpyxl worksheet containing the source data
        - prefix (str): prefix for naming the new sheet

        Returns
        - Worksheet: the newly created worksheet with the chart
        """
        chart = self.create_line_chart(
            "Throughput Variations in Records / Seconds",
            INTERVALS_AXIS_TITLE,
            "Throughput in Records/Sec",
        )
        # adding data
        chart.append(self.get_throughput_write_request_records_series(ws, prefix))
        chart.append(self.get_throughput_read_request_records_series(ws, prefix))
        chart.append(self.get_throughput_records_series(ws, prefix))
        # add chart to the sheet
        sheet = self.wb.create_sheet("Records_Sec")
        sheet.add_chart(chart)
        return sheet

    def ensure_sbk_logo(self, img_path=None, cell='K7', scale=0.5):
        """Insert or verify the presence of the SBK logo in the SBK sheet.

        Parameters
        - img_path (str): path to the SBK logo image file
        - cell (str): Excel cell address where the image should be placed
        - scale (float): scaling factor for the image size

        Returns
        - None

        Notes
        - If the image already exists in the specified cell, it will not be
          re-inserted.
        - The image is loaded from the local file system; ensure the path is
          correct.
        """
        ws = self.wb[sheets_constants.SBK_SHEET]
        # Put your sheet in the loader
        image_loader = SheetImageLoader(ws)

        if image_loader.image_in(cell):
            print(f"SBK logo Image already exists in SBK sheet")
            return

        if img_path is None:
            img_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                    'images', 'sbk-logo.png')
        # Check if image already exists in the sheet
        img_abs_path = os.path.abspath(img_path)
        # Add image if not present
        if os.path.exists(img_abs_path):
            img = Image(img_abs_path)
            img.width *= scale
            img.height *= scale
            ws.add_image(img, cell)
            print(f"Inserted image at {img_abs_path} in SBK sheet")
        else:
            print(f"Image not found: {img_abs_path}")

    def create_graphs(self):
        """Top-level method to create all graphs for the SBK report.

        This will:
        - create throughput and latency graphs for the R sheet
        - create total latency percentile graphs for the T sheet
        - save the modified workbook to the original file

        Parameters
        - None

        Returns
        - None

        Notes
        - The method assumes the first row of each sheet is a header row.
        - Existing sheets are not deleted; new charts are added to the next
          available sheet.
        """
        r_name = sheets_constants.FIRST_RESULT_SHEET
        r_ws = self.wb[r_name]
        r_prefix = r_name + get_storage_name_from_worksheet(r_ws)
        t_name = sheets_constants.FIRST_TOTAL_SHEET
        t_ws = self.wb[t_name]
        t_prefix = t_name + get_storage_name_from_worksheet(t_ws)
        self.create_throughput_mb_graph(r_ws, r_prefix)
        self.create_throughput_records_graph(r_ws, r_prefix)
        self.create_latency_compare_graphs(r_ws, r_prefix)
        self.create_latency_graphs(r_ws, r_prefix)
        self.create_total_latency_percentile_graphs(t_ws, t_prefix)
        self.apply_table_theme()
        self.apply_chart_theme()
        self.wb.save(self.file)
        print("file : %s updated with graphs" % self.file)
